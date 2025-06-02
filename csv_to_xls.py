import pandas as pd
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def csv_to_xls(csv_file, output_file=None):
    """
    Convert CSV file to XLS format with specific columns required by the template.

    Args:
        csv_file (str): Path to the CSV file to convert
        output_file (str, optional): Output XLS file path. If None, uses the CSV filename with .xls extension.
    """
    if not output_file:
        output_file = os.path.splitext(csv_file)[0] + '.xls'

    # Read the CSV file
    print(f"Reading CSV file: {csv_file}")
    df = pd.read_csv(csv_file)

    # Create a new workbook and get the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"

    # Define the headers for the XLS template
    headers = [
        'Referência (código fornecedor)',
        'Código do produto (ID Tray)',
        'Nome do produto',
        'Preço de venda em reais',
        'Preço de custo em reais',
        'Estoque do produto',
        'Exibir produto ativo',
        'Prazo de disponibilidade',
        'SEO - Endereço do produto (URL)',
        'SEO - Palavras chaves do produto',
        'Peso do produto (gramas)',
        'HTML da descrição completa'
    ]

    # Set column styles
    header_font = Font(bold=True)
    header_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    header_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers to the first row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border

    # Map CSV columns to XLS template columns with fallbacks
    primary_mapping = {
        'Referência (código fornecedor)': 'Handle',
        'Código do produto (ID Tray)': 'Variant SKU',
        'Nome do produto': 'Title',
        'Preço de venda em reais': 'Variant Price',
        'Preço de custo em reais': 'Cost per item',
        'Estoque do produto': None,  # We'll handle stock separately to ensure numeric values
        'Exibir produto ativo': 'Published',
        'Prazo de disponibilidade': None,
        'SEO - Endereço do produto (URL)': 'Handle',
        'SEO - Palavras chaves do produto': 'Tags',
        'Peso do produto (gramas)': 'Variant Grams',
        'HTML da descrição completa': 'Body (HTML)',
    }

    # Secondary mappings for fallbacks
    secondary_mapping = {
        'Código do produto (ID Tray)': ['Product ID', 'SKU', 'ID'],
        'Nome do produto': ['Product Title', 'Product Name'],
        'Preço de venda em reais': ['Price', 'Sale Price', 'Variant Compare At Price'],
        'Peso do produto (gramas)': ['Weight', 'Weight (g)', 'Product Weight'],
        'HTML da descrição completa': ['Description', 'Product Description', 'HTML Description'],
    }

    # Process data rows
    print(f"Processing {len(df)} products...")
    row_idx = 2  # Start from second row (after header)
    error_count = 0

    for _, row in df.iterrows():
        for col_idx, header in enumerate(headers, 1):
            value = None

            # Handle stock quantity separately to ensure it's numeric
            if header == 'Estoque do produto':
                # Try to get stock quantity directly from one of these fields
                stock_fields = ['Variant Inventory Qty', 'Quantity', 'Stock', 'Inventory']
                stock_value = None

                for field in stock_fields:
                    if field in row and not pd.isna(row[field]):
                        try:
                            stock_value = int(float(row[field]))
                            break
                        except (ValueError, TypeError):
                            continue

                # If we couldn't find a valid stock value, determine based on inventory policy
                if stock_value is None:
                        stock_value = 0

                # Ensure it's a valid integer
                value = max(0, stock_value)  # Make sure it's not negative
            else:
                # Try primary mapping first
                primary_column = primary_mapping.get(header)
                if primary_column and primary_column in row:
                    value = row[primary_column]

                # Try secondary mappings if value is None or empty
                if (value is None or pd.isna(value) or value == '') and header in secondary_mapping:
                    for fallback_column in secondary_mapping[header]:
                        if fallback_column in row and not pd.isna(row[fallback_column]) and row[fallback_column] != '':
                            value = row[fallback_column]
                            break

            # Special handling for specific columns
            if header == 'Exibir produto ativo':
                # Convert to "S" for true and "N" for false
                if value is not None:
                    value = 'S' if str(value).lower() == 'true' else 'N'
                else:
                    value = 'S'  # Default to active

            elif header == 'Preço de venda em reais' or header == 'Preço de custo em reais':
                # Ensure price is formatted properly
                try:
                    value = float(value) if value is not None and pd.notna(value) else 0.0
                except (ValueError, TypeError):
                    value = 0.0

            elif header == 'SEO - Endereço do produto (URL)':
                # Use Handle or construct a URL from the product name
                if value is None or pd.isna(value) or value == '':
                    if 'Title' in row and not pd.isna(row['Title']):
                        # Create a URL-friendly version of the title
                        title = str(row['Title']).lower()
                        title = title.replace(' ', '-').replace('/', '-').replace('_', '-')
                        title = ''.join(c for c in title if c.isalnum() or c == '-')
                        value = title

            elif header == 'Prazo de disponibilidade':
                # Use availability data if available, otherwise default value
                if 'Availability' in row:
                    value = row['Availability']
                else:
                    value = '5'  # Default to 5 days

            elif header == 'Peso do produto (gramas)':
                # Ensure weight is in grams
                if value is not None and pd.notna(value):
                    try:
                        weight = float(value)
                        # Check if weight seems to be in kg (less than 10)
                        if weight < 10 and 'Variant Weight Unit' in row:
                            weight_unit = str(row['Variant Weight Unit']).lower()
                            if weight_unit == 'kg':
                                weight *= 1000  # Convert kg to g
                        value = weight
                    except (ValueError, TypeError):
                        value = 0

            # Validation for Nome do produto (must not be empty)
            elif header == 'Nome do produto':
                if value is None or pd.isna(value) or str(value).strip() == '':
                    # Use a default value if product name is empty
                    if 'Handle' in row and not pd.isna(row['Handle']):
                        value = f"Product {str(row['Handle'])}"
                    else:
                        value = f"Product {row_idx-1}"
                    print(f"Warning: Empty product name in row {row_idx}, using '{value}' instead")

            # Validation for Código do produto (ID Tray) (must be numeric)
            elif header == 'Código do produto (ID Tray)':
                if value is not None and pd.notna(value):
                    # Try to convert to numeric
                    try:
                        value = int(float(str(value).replace(',', '.').strip()))
                    except (ValueError, TypeError):
                        # If conversion fails, use row index as a numeric ID
                        value = row_idx + 1000  # Adding 1000 to avoid low numbers
                        print(f"Warning: Non-numeric ID in row {row_idx}, using {value} instead")
                else:
                    # If empty, use row index as default numeric ID
                    value = row_idx + 1000
                    print(f"Warning: Missing ID in row {row_idx}, using {value} instead")

            # Set the cell value
            ws.cell(row=row_idx, column=col_idx, value=value)

        row_idx += 1

    # Auto-adjust column widths for better readability
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Add some padding and set a minimum width
        adjusted_width = max(12, max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the file
    print(f"Saving to {output_file}")
    wb.save(output_file)
    print(f"Conversion complete. Output saved to {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python csv_to_xls.py input.csv [output.xls]")
        sys.exit(1)

    csv_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None

    csv_to_xls(csv_file, output_file)